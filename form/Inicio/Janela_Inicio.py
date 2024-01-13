import sys
from datetime import datetime, timedelta
from PyQt6.QtWidgets import QMainWindow, QApplication
from PyQt6 import QtWidgets, QtCore
import pandas as pd
from openpyxl import Workbook
from PyQt6.QtCore import Qt
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter.filedialog
from form.Inicio.Form_Inicio import Ui_Form_Inicio
from funcoes.Banco.Conexao_banco import Classe_Banco
from funcoes.Alertas.Arquivo_Alertas import Classe_Alertas


class Classe_Inicio(QMainWindow, Ui_Form_Inicio):
    def __init__(self):
        super(Classe_Inicio, self).__init__()
        self.setupUi(self)
        self.banco = Classe_Banco()
        self.alert = Classe_Alertas()
    

        self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        self.Tx_cliente_relatorio.textChanged.connect(self.pesquisa_cliente_relatorio)
        self.Tx_Usuario_relatorio.textChanged.connect(self.pesquisa_operador_relatorio)
        self.Tx_Venda_relatorio.textChanged.connect(self.pesquisa_n_venda_relatorio)
        self.Bt_Ft_Pagamento.clicked.connect(self.filtro_forma_pgmt_relatorio)
        self.Bt_FtData.clicked.connect(self.filtro_data_relatorio)
        self.tx_BuscaProdutos.textChanged.connect(self.pesquisar_produtos)
        self.tx_BuscaClientes.textChanged.connect(self.pesquisar_clientes)
        self.Bt_Excluir_Cliente.clicked.connect(self.excluir_clientes)
        self.Bt_Remover_Usuario.clicked.connect(self.excluir_usuarios)
        self.Bt_Excluir_Produto.clicked.connect(self.excluir_produtos)
        self.Bt_Sair.clicked.connect(self.fechar_tela_inicio)
        self.Bt_Imprimir_Atendimento.clicked.connect(self.imprimir_atendimento)
        self.tx_BuscaUsuarios.textChanged.connect(self.pesquisar_usuarios)
        self.Bt_Max_Jan.clicked.connect(self.maximizar_jan)
        self.Bt_Min_Jan.clicked.connect(self.minimizar_jan)
        self.Bt_Fechar_Jan.clicked.connect(self.fechar_tela_inicio)


    ######## --- Chama StakeWidgets --- ########
        self.Bt_Inicio.clicked.connect(self.tela_inicio)
        self.Bt_Produtos.clicked.connect(self.tela_produtos)
        self.Bt_Clientes.clicked.connect(self.tela_clientes)
        self.Bt_Relatorio.clicked.connect(self.tela_relatorio)
        self.Bt_Fornecedores.clicked.connect(self.tela_fornecedores)
        self.Bt_Usuarios.clicked.connect(self.tela_usuarios)
        self.Bt_Suporte.clicked.connect(self.tela_suporte)
        self.frame_lateral.enterEvent = lambda event: self.expaandir_left_menu()
        self.frame_lateral.leaveEvent  = lambda event: self.expaandir_left_menu()
        
        self.venda_hj()    
        self.vendas_mes_atual()
        
    ##### --- Função de Permissões de acesso --- #####
    def permissoes_visualizar(self):
        self.Bt_Add_Fornecedor.setDisabled(True)
        self.Bt_Remover_Fornecedor.setDisabled(True)
        self.Bt_Edit_Fornecedor.setDisabled(True)
        self.Bt_Add_Fornecedor.setDisabled(True)
        self.Bt_Excluir_Produto.setDisabled(True)
        self.Bt_Edit_Produto.setDisabled(True)
        self.Bt_Add_Produto.setDisabled(True)
        self.Bt_Excluir_Cliente.setDisabled(True)
        self.Bt_Edit_Cliente.setDisabled(True)
        self.Bt_Cad_Cliente.setDisabled(True)
        self.Bt_Usuarios.setDisabled(True)

    def tela_inicio(self):
        self.stackedWidget.setCurrentIndex(0)
        self.venda_hj() 
        self.vendas_mes_atual()

    def tela_produtos(self):
        self.stackedWidget.setCurrentIndex(5)
        self.listar_produtos()

    def tela_clientes(self):
        self.stackedWidget.setCurrentIndex(4)
        self.listar_clientes()

    def tela_relatorio(self):
        self.stackedWidget.setCurrentIndex(2)
        self.listar_relatorio()

    def tela_fornecedores(self):
        self.stackedWidget.setCurrentIndex(6)

    
    def tela_usuarios(self):
        self.stackedWidget.setCurrentIndex(7)
        self.listar_usuarios()

    def tela_suporte(self):
        self.stackedWidget.setCurrentIndex(8)


    ##### --- Função de Permissões de acesso --- #####
    def permissoes_visualizar(self):
        self.Bt_Add_Fornecedor.setDisabled(True)
        self.Bt_Remover_Fornecedor.setDisabled(True)
        self.Bt_Edit_Fornecedor.setDisabled(True)
        self.Bt_Add_Fornecedor.setDisabled(True)
        self.Bt_Excluir_Produto.setDisabled(True)
        self.Bt_Edit_Produto.setDisabled(True)
        self.Bt_Add_Produto.setDisabled(True)
        self.Bt_Excluir_Cliente.setDisabled(True)
        self.Bt_Edit_Cliente.setDisabled(True)
        self.Bt_Cad_Cliente.setDisabled(True)
        self.Bt_Usuarios.setDisabled(True)


    def minimizar_jan(self):
        self.showMinimized()


    def maximizar_jan(self):
        self.showMaximized()

    
    def focus_codigo(self):
        self.Input_Codigo.setFocus()


    def pesquisar_clientes(self):
        resultados = []
        valor_consulta = self.tx_BuscaClientes.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.clientes")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if valor_consulta.lower() in pessoa[1].lower():
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Cliente.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Cliente.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Cliente.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.close()
        self.banco.cursorr.close()

        
########## --- FITLTRO  RELATÓRIO VENDAS --- ########## 
    def pesquisa_cliente_relatorio(self):
        resultados = []
        consulta_cliente_relatorio = self.Tx_cliente_relatorio.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if consulta_cliente_relatorio.lower() in pessoa[5].lower():
                resultados.append(pessoa)
        if not dados_pessoais:
            return  self.alert.alerta_registro()
        else:   
            self.TableWidget_Relatorio.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Relatorio.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Relatorio.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.close()
        self.banco.cursorr.close()

    
    def pesquisa_n_venda_relatorio(self):
        consulta_n_venda_relatorio = self.Tx_Venda_relatorio.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas WHERE idvendas LIKE '%{consulta_n_venda_relatorio}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alert.alerta_registro()    
        else:   
            self.TableWidget_Relatorio.setRowCount(0)
            for idxLinha, linha in enumerate(lista):
                self.TableWidget_Relatorio.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Relatorio.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
        self.banco.cursorr.close()
    

    def filtro_forma_pgmt_relatorio(self):
        resultados = []
        consulta_frm_pgmt_relatorio = self.Tx_FiltroStatus.currentText()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if consulta_frm_pgmt_relatorio.lower() in pessoa[4].lower():
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Relatorio.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Relatorio.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Relatorio.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.close()
        self.banco.cursorr.close()


    def filtro_data_relatorio(self):
        resultados = []
        consulta_data_relatorio = self.Tx_FiltroData.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if consulta_data_relatorio in pessoa[1]:
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Relatorio.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Relatorio.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Relatorio.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.close()
        self.banco.cursorr.close()
    

    def pesquisa_operador_relatorio(self):
        resultados = []
        consulta_operador_relatorio = self.Tx_Usuario_relatorio.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if consulta_operador_relatorio in pessoa[3]:
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Relatorio.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Relatorio.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Relatorio.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
        self.banco.cursorr.close()
    

    def pesquisar_produtos(self):
        resultados = []
        valor_consulta = self.tx_BuscaProdutos.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.produtos")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if valor_consulta.lower() in pessoa[1].lower():
                resultados.append(pessoa)
            elif valor_consulta in pessoa[5]:
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Produto.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Produto.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Produto.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.close()
        self.banco.cursorr.close()
    

    def pesquisar_usuarios(self):
        resultados = []
        valor_consulta = self.tx_BuscaUsuarios.text()
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.usuarios")
        dados_pessoais = self.banco.cursorr.fetchall()
        for pessoa in dados_pessoais:
            if valor_consulta.lower() in pessoa[1].lower():
                resultados.append(pessoa)
            elif valor_consulta.lower() in pessoa[2].lower():
                resultados.append(pessoa)
        if not resultados:
            return  self.alert.alerta_registro()     
        else:   
            self.TableWidget_Usuario.setRowCount(0)
            for idxLinha, linha in enumerate(resultados):
                self.TableWidget_Usuario.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.TableWidget_Usuario.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
       

    ######## --- Funções Listar listar Itens --- ######## 
    def listar_clientes(self):
        self.TableWidget_Cliente.verticalHeader().hide()
        col_widths = [50, 200, 100, 90, 90, 150, 77, 200, 60, 200, 150, 50, 100, 100]
        for i, width in enumerate(col_widths):
            self.TableWidget_Cliente.setColumnWidth(i, width)
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT * FROM pdv.clientes")
        dados_lidosc = self.banco.cursorr.fetchall()
        self.TableWidget_Cliente.setRowCount(len(dados_lidosc))
        self.TableWidget_Cliente.setColumnCount(15)
        for a, dados in enumerate(dados_lidosc):
            for b, valor in enumerate(dados):
                item = QtWidgets.QTableWidgetItem(str(valor))
                self.TableWidget_Cliente.setItem(a, b, item)
        self.banco.query.commit()
        self.banco.cursorr.close()

        
    def listar_relatorio(self):
        self.TableWidget_Relatorio.verticalHeader().hide()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT * FROM pdv.vendas")
        dados_lidos = self.banco.cursorr.fetchall()
        self.TableWidget_Relatorio.setRowCount(len(dados_lidos))
        self.TableWidget_Relatorio.setColumnCount(6)
        for a, dados in enumerate(dados_lidos):
            for b, valor in enumerate(dados):
                item = QtWidgets.QTableWidgetItem(str(valor))
                self.TableWidget_Relatorio.setItem(a, b, item)
        self.banco.query.commit()
        self.banco.cursorr.close()

    
    def listar_produtos(self):
        self.TableWidget_Produto.verticalHeader().hide()
        col_widths = [50, 200, 100, 90, 70, 90, 70, 70, 100, 100, 100]
        for i, width in enumerate(col_widths):
            self.TableWidget_Produto.setColumnWidth(i, width)
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT * FROM pdv.produtos")
        dados_lidos = self.banco.cursorr.fetchall()
        self.TableWidget_Produto.setRowCount(len(dados_lidos))
        self.TableWidget_Produto.setColumnCount(11)
        for a, dados in enumerate(dados_lidos):
            for b, valor in enumerate(dados):
                item = QtWidgets.QTableWidgetItem(str(valor))
                self.TableWidget_Produto.setItem(a, b, item)
        self.banco.query.commit()
        self.banco.cursorr.close()

    
    def listar_usuarios(self):
        self.TableWidget_Usuario.verticalHeader().hide()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idusuarios, nome, login, nivel_de_acesso, permissao FROM pdv.usuarios")
        dados_lidos = self.banco.cursorr.fetchall()
        self.TableWidget_Usuario.setRowCount(len(dados_lidos))
        self.TableWidget_Usuario.setColumnCount(5)
        for a, dados in enumerate(dados_lidos):
            for b, valor in enumerate(dados):
                item = QtWidgets.QTableWidgetItem(str(valor))
                self.TableWidget_Usuario.setItem(a, b, item)
        self.banco.query.commit()
        self.banco.cursorr.close()
        self.banco.query.close()
    

    ######## --- Funções Remover Itens --- ######## 
    def excluir_usuarios(self):
        selected_row = self.TableWidget_Usuario.currentRow()
        if selected_row >= 0:
            self.id_usuarios = int(self.TableWidget_Usuario.item(selected_row, 0).text())
            self.TableWidget_Usuario.removeRow(selected_row)
            self.banco.conectar()
            delete_query = f"DELETE FROM pdv.usuarios WHERE idusuarios = {self.id_usuarios}"
            self.banco.cursorr.execute(delete_query)
            self.banco.query.commit()
            self.banco.cursorr.close()
            self.banco.query.close()
            self.tx_BuscaUsuarios.clear()
            self.listar_usuarios()

    
    def excluir_clientes(self):
        selected_row = self.TableWidget_Cliente.currentRow()
        if selected_row >= 0:
            self.id_clientes = int(self.TableWidget_Cliente.item(selected_row, 0).text())
            self.TableWidget_Cliente.removeRow(selected_row)
            self.banco.conectar()
            delete_query = f"DELETE FROM pdv.clientes WHERE idclientes = {self.id_clientes}"
            self.banco.cursorr.execute(delete_query)
            self.banco.query.commit()
            self.banco.cursorr.close()
            self.banco.query.close()
            self.tx_BuscaClientes.clear()
            self.listar_clientes()
    

    def excluir_produtos(self):
        selected_row = self.TableWidget_Produto.currentRow()
        if selected_row >= 0:
            self.id_produtos = int(self.TableWidget_Produto.item(selected_row, 0).text())
            self.TableWidget_Produto.removeRow(selected_row)
            self.banco.conectar()
            delete_query = f"DELETE FROM pdv.produtos WHERE idprodutos = {self.id_produtos}"
            self.banco.cursorr.execute(delete_query)
            self.banco.query.commit()
            self.banco.cursorr.close()
            self.banco.query.close()
            self.tx_BuscaProdutos.clear()
            self.listar_produtos()


    ##### --- Função de  Expandir o menu lateral --- ######
    def expaandir_left_menu(self):
        tamanho = self.frame_lateral.width()
        if tamanho == 100:
            novo_tamanho = 50
            self.Bt_Clientes.setText("")
            self.Bt_Mesas.setText("")
            self.Bt_Inicio.setText("")
            self.Bt_Produtos.setText("")
            self.Bt_Sair.setText("")
            self.Bt_Suporte.setText("")
            self.Bt_Fornecedores.setText("")
            self.Bt_Usuarios.setText("")
            self.Bt_Relatorio.setText("")
            self.Bt_Vendas.setText("")
        else:
            novo_tamanho = 100
            self.Bt_Inicio.setText('Inicio')
            self.Bt_Clientes.setText("Clientes")
            self.Bt_Fornecedores.setText("Fornecedores")
            self.Bt_Produtos.setText("Produtos")
            self.Bt_Mesas.setText("Mesas")  
            self.Bt_Vendas.setText("Pdv")
            self.Bt_Usuarios.setText("Usuários")
            self.Bt_Relatorio.setText("Relatório")
            self.Bt_Suporte.setText("Suporte")
            self.Bt_Sair.setText("Sair")
        self.animacao = QtCore.QPropertyAnimation(self.frame_lateral,  b"maximumWidth")
        self.animacao.setStartValue(tamanho)
        self.animacao.setEndValue(novo_tamanho)
        self.animacao.setDuration(390)
        self.animacao.start()  

    
    def focus_codigo(self):
        self.Input_Codigo.setFocus()

    
    def imprimir_atendimento(self):
        try:
            data = {}
            headers = []
            for col in range(self.TableWidget_Relatorio.columnCount()):
                header_text = self.TableWidget_Relatorio.horizontalHeaderItem(col).text()
                headers.append(header_text)
                data[header_text] = [self.TableWidget_Relatorio.item(row, col).text() for row in range(self.TableWidget_Relatorio.rowCount())]
            df = pd.DataFrame.from_dict(data)
            local_salvar = tkinter.filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)  #### --- Adiciona o cabeçalho à planilha --- ###
            for row in df.itertuples(index=False):
                sheet.append(row)
            for col_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = adjusted_width
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fill_type='solid', fgColor='AAAAAA')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            workbook.save(local_salvar)
            self.alert.alerta_salvar_relatorio()
        except:
            self.alert.alerta_erro_salvar_relatorio()


    def fechar_tela_inicio(self):
        self.close() 

######### --- DASHBOARD - INICIO --- #########
    def venda_hj(self):
        soma_valor_venda = 0
        data_atual = datetime.now()
        data_formatada = data_atual.strftime("%d/%m/%Y")
        self.banco.conectar()
        self.banco.cursorr.execute(f"SELECT * FROM pdv.vendas WHERE data LIKE '%{data_formatada}%'")
        resultados = self.banco.cursorr.fetchall()
        self.banco.cursorr.close()
        self.banco.query.close()
        for resultado in resultados:
            valor_venda = float(resultado[2])
            soma_valor_venda += valor_venda
        self.Lb_Vendas_Hj.setText(str(f'{soma_valor_venda:.2f}'))


    def vendas_mes_atual(self):
        soma_valor_venda_mes = 0
        data_atual = datetime.now()
        primeiro_dia_do_mes = data_atual.replace(day=1)
        ultimo_dia_do_mes = (data_atual.replace(month=data_atual.month + 1, day=1) - timedelta(days=1)).replace(hour=23, minute=59, second=59)
        data_inicial_formatada = primeiro_dia_do_mes.strftime('%d/%m/%Y')
        data_final_formatada = ultimo_dia_do_mes.strftime('%d/%m/%Y')
        self.banco.conectar()
        consulta_sql = "SELECT * FROM pdv.vendas WHERE STR_TO_DATE(data, '%d/%m/%Y') BETWEEN STR_TO_DATE(%s, '%d/%m/%Y') AND STR_TO_DATE(%s, '%d/%m/%Y')"
        parametros = (data_inicial_formatada, data_final_formatada)
        self.banco.cursorr.execute(consulta_sql, parametros)
        resultados = self.banco.cursorr.fetchall()
        self.banco.cursorr.close()
        self.banco.query.close()
        for resultado in resultados:
            valor_venda = float(resultado[2])
            soma_valor_venda_mes += valor_venda
            self.Lb_Venda_Mes.setText(str(f"{soma_valor_venda_mes:.2f}"))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    systen = Classe_Inicio()
    systen.showMaximized()
    sys.exit(app.exec())